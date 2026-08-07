from datetime import datetime

from openpyxl import load_workbook


from config import (
    SHEET_GAMMA_SLIDES,
    SHEET_DESCRIPTIONS,
    SHEET_ASSET_REGISTER,
    SHEET_BUILD_LOG
)


# ==========================================================
# Presentation Writer
# ==========================================================

class PresentationWriter:

    def __init__(

            self,

            workbook_path

    ):

        self.workbook_path = workbook_path

        self.workbook = load_workbook(

            workbook_path

        )

    # ======================================================
    # Header Map
    # ======================================================

    @staticmethod
    def header_map(sheet):

        headers = {}

        for cell in sheet[1]:

            if cell.value:
                headers[

                    str(cell.value).strip()

                ] = cell.column

        return headers

    # ======================================================
    # Gamma Slides
    # ======================================================

    def update_gamma_slides(

            self,

            lesson_package_id,

            deck_id,

            slides_id,

            slides_url,

            gamma_embed_url,

            slide_title,

            slide_number=1,

            speaker_notes="",
    ):

        sheet = self.workbook[

            SHEET_GAMMA_SLIDES

        ]

        headers = self.header_map(

            sheet

        )

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:
                break

            if value == lesson_package_id:

                #
                # Slide Title
                #
                if "slide_title" in headers:
                    sheet.cell(

                        row=row,

                        column=headers["slide_title"]

                    ).value = slide_title

                #
                # Slide Number
                #

                sheet.cell(

                    row=row,

                    column=headers["slide_number"]

                ).value = slide_number

                #
                # Speaker Notes
                #

                sheet.cell(

                    row=row,

                    column=headers["speaker_notes"]

                ).value = speaker_notes

                #
                # Gamma Deck
                #

                sheet.cell(

                    row=row,

                    column=headers["gamma_deck_id"]

                ).value = deck_id

                #
                # Gamma Slides
                #

                sheet.cell(

                    row=row,

                    column=headers["gamma_slides_id"]

                ).value = slides_id

                sheet.cell(

                    row=row,

                    column=headers["gamma_slides_url"]

                ).value = slides_url

                #
                # Gamma Embed URL
                #

                sheet.cell(

                    row=row,

                    column=headers["gamma_embed_url"]

                ).value = gamma_embed_url

                #
                # Embed HTML
                #

                iframe = (

                    f'<iframe '

                    f'src="{gamma_embed_url}" '

                    f'width="100%" '

                    f'height="720" '

                    f'allowfullscreen>'

                    f'</iframe>'

                )

                sheet.cell(

                    row=row,

                    column=headers["slides_embed_html"]

                ).value = iframe

                #
                # Status
                #

                sheet.cell(

                    row=row,

                    column=headers["status"]

                ).value = "COMPLETED"

                return

            row += 1

            # ======================================================
            # Descriptions Worksheet
            # ======================================================

    def update_descriptions(

                self,

                lesson_package_id,

                slides_title,

                slides_description,

                generation_status,

                review_status

        ):

            sheet = self.workbook[

                SHEET_DESCRIPTIONS

            ]
            print("=" * 60)
            print("UPDATING DESCRIPTIONS")
            print("Lesson Package:", lesson_package_id)
            print("=" * 60)

            headers = self.header_map(

                sheet

            )

            row = 2

            while True:

                value = sheet.cell(

                    row=row,

                    column=headers["lesson_package_id"]

                ).value
                print("Workbook row:", row)
                print("Lesson package:", value)
                if not value:
                    break
                print("Workbook row:", row, "Lesson:", value)
                if value == lesson_package_id:
                    print("MATCH FOUND")
                    if "slides_title" in headers:
                        sheet.cell(

                            row=row,

                            column=headers["slides_title"]

                        ).value = slides_title

                    if "slides_description" in headers:
                        sheet.cell(

                            row=row,

                            column=headers["slides_description"]

                        ).value = slides_description

                    if "generation_status" in headers:
                        sheet.cell(

                            row=row,

                            column=headers["generation_status"]

                        ).value = generation_status

                    if "review_status" in headers:
                        sheet.cell(

                            row=row,

                            column=headers["review_status"]

                        ).value = review_status
                    print("DESCRIPTION SAVED")
                    return

                row += 1

            print("=" * 60)
            print("DESCRIPTION HEADERS")
            print(headers)
            print("=" * 60)
        # ======================================================
        # Asset Register
        # ======================================================

    def update_asset_register(

            self,

            lesson_package_id,

            asset_type,

            filename,

            url,

            status="COMPLETED"

    ):

        sheet = self.workbook[

            SHEET_ASSET_REGISTER

        ]

        headers = self.header_map(

            sheet

        )

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:
                break

            current_type = sheet.cell(

                row=row,

                column=headers["asset_type"]

            ).value

            if (

                    value == lesson_package_id

                    and

                    current_type == asset_type

            ):
                #
                # Filename
                #

                sheet.cell(

                    row=row,

                    column=headers["asset_filename"]

                ).value = filename

                #
                # URL
                #

                sheet.cell(

                    row=row,

                    column=headers["asset_url"]

                ).value = url

                #
                # Status
                #

                sheet.cell(

                    row=row,

                    column=headers["status"]

                ).value = status

                return

            row += 1

        #
        # Asset not found
        # Create a new row
        #

        new_row = sheet.max_row + 1

        sheet.cell(

            row=new_row,

            column=headers["lesson_package_id"]

        ).value = lesson_package_id

        sheet.cell(

            row=new_row,

            column=headers["asset_type"]

        ).value = asset_type

        sheet.cell(

            row=new_row,

            column=headers["asset_filename"]

        ).value = filename

        sheet.cell(

            row=new_row,

            column=headers["asset_url"]

        ).value = url

        sheet.cell(

            row=new_row,

            column=headers["status"]

        ).value = status

    # ======================================================
    # Build Log
    # ======================================================

    def log(

            self,

            component,

            action,

            status,

            details

    ):

        sheet = self.workbook[

            SHEET_BUILD_LOG

        ]

        sheet.append([

            datetime.now().strftime(

                "%Y-%m-%d %H:%M:%S"

            ),

            component,

            action,

            status,

            details

        ])

    # ======================================================
    # Save Workbook
    # ======================================================

    def save(self):

        self.workbook.save(

            self.workbook_path

        )

        self.workbook.close()
