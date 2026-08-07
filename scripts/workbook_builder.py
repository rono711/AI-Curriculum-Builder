from copy import copy

from openpyxl import load_workbook

from config import (
    TEMPLATE_WORKBOOK,
    OUTPUT_WORKBOOK
)


class WorkbookBuilder:

    def __init__(self):

        self.workbook = load_workbook(TEMPLATE_WORKBOOK)

    # ----------------------------------------------------

    def worksheet(self, name):

        return self.workbook[name]

    # ----------------------------------------------------

    def clear_sheet(self, sheet_name):

        ws = self.worksheet(sheet_name)

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

    # ----------------------------------------------------

    def write_rows(self, sheet_name, rows):

        ws = self.worksheet(sheet_name)

        self.clear_sheet(sheet_name)

        if len(rows) == 0:
            return

        headers = list(rows[0].keys())

        for row in rows:

            values = []

            for h in headers:
                values.append(row.get(h, ""))

            ws.append(values)

    # ----------------------------------------------------

    def build(self, workbook_request):

        worksheets = workbook_request["worksheets"]

        self.write_rows(
            "01_Curriculum_Master",
            worksheets["Curriculum_Master"]
        )

        self.write_rows(
            "02_AI_Generation",
            worksheets["AI_Generation"]
        )

        self.write_rows(
            "03_Prompt_Library",
            worksheets["Prompt_Library"]
        )

        self.write_rows(
            "04_Moodle_Mapping",
            worksheets["Moodle_Mapping"]
        )

        self.write_rows(
            "05_Generation_Log",
            worksheets["Generation_Log"]
        )

        self.workbook.save(OUTPUT_WORKBOOK)

        return {
            "status": "success",
            "filename": str(OUTPUT_WORKBOOK)
        }
