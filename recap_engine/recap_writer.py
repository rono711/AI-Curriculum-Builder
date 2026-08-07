from datetime import datetime

from openpyxl import load_workbook

from config import (
    SHEET_RECAP,
    SHEET_DESCRIPTIONS,
    SHEET_ASSET_REGISTER,
    SHEET_BUILD_LOG
)


# ==========================================================
# Recap Writer
# ==========================================================

class RecapWriter:

    def __init__(

            self,

            workbook_path

    ):

        self.workbook_path = workbook_path

        self.workbook = load_workbook(

            workbook_path

        )

    # ======================================================
    # Header Map
    # ======================================================

    @staticmethod
    def header_map(sheet):

        headers = {}

        for cell in sheet[1]:

            if cell.value:

                headers[

                    str(cell.value).strip()

                ] = cell.column

        return headers

    # ======================================================
    # Recap Worksheet
    # ======================================================

    def update_recap(

            self,

            lesson_package_id,

            markdown_filename,

            html_filename,

            generation_status,

            review_status

    ):

        sheet = self.workbook[

            SHEET_RECAP

        ]

        headers = self.header_map(

            sheet

        )

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:

                break

            if value == lesson_package_id:

                if "recap_markdown" in headers:

                    sheet.cell(

                        row=row,

                        column=headers["recap_markdown"]

                    ).value = markdown_filename

                if "recap_html" in headers:

                    sheet.cell(

                        row=row,

                        column=headers["recap_html"]

                    ).value = html_filename

                if "generation_status" in headers:

                    sheet.cell(

                        row=row,

                        column=headers["generation_status"]

                    ).value = generation_status

                if "review_status" in headers:

                    sheet.cell(

                        row=row,

                        column=headers["review_status"]

                    ).value = review_status

                return

            row += 1
 # ======================================================
 # Descriptions Worksheet
 # ======================================================

    def update_descriptions(

            self,

            lesson_package_id,

            recap_title,

            recap_description,

            generation_status,

            review_status

    ):

        sheet = self.workbook[

            SHEET_DESCRIPTIONS

        ]
        print("=" * 60)
        print("UPDATING DESCRIPTIONS")
        print("Lesson Package:", lesson_package_id)
        print("=" * 60)

        headers = self.header_map(

            sheet

        )

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:
                break
            print("Workbook row:", row, "Lesson:", value)
            if value == lesson_package_id:
                print("MATCH FOUND")
                if "recap_title" in headers:
                    sheet.cell(

                        row=row,

                        column=headers["recap_title"]

                    ).value = recap_title

                if "recap_description" in headers:
                    sheet.cell(

                        row=row,

                        column=headers["recap_description"]

                    ).value = recap_description

                if "generation_status" in headers:
                    sheet.cell(

                        row=row,

                        column=headers["generation_status"]

                    ).value = generation_status

                if "review_status" in headers:
                    sheet.cell(

                        row=row,

                        column=headers["review_status"]

                    ).value = review_status
                print("DESCRIPTION SAVED")
                return

            row += 1
    # ======================================================
    # Asset Register
    # ======================================================

    def update_asset_register(

            self,

            lesson_package_id,

            asset_type,

            filename,

            url,

            status="COMPLETED"

    ):

        sheet = self.workbook[

            SHEET_ASSET_REGISTER

        ]

        headers = self.header_map(

            sheet

        )

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:

                break

            current = sheet.cell(

                row=row,

                column=headers["asset_type"]

            ).value

            if (

                value == lesson_package_id

                and

                current == asset_type

            ):

                sheet.cell(

                    row=row,

                    column=headers["asset_filename"]

                ).value = filename

                sheet.cell(

                    row=row,

                    column=headers["asset_url"]

                ).value = url

                if "status" in headers:

                    sheet.cell(

                        row=row,

                        column=headers["status"]

                    ).value = status

                return

            row += 1

    # ======================================================
    # Build Log
    # ======================================================

    def log(

            self,

            component,

            action,

            status,

            details

    ):

        sheet = self.workbook[

            SHEET_BUILD_LOG

        ]

        sheet.append([

            datetime.now().strftime(

                "%Y-%m-%d %H:%M:%S"

            ),

            component,

            action,

            status,

            details

        ])

    # ======================================================
    # Save
    # ======================================================

    def save(self):

        self.workbook.save(

            self.workbook_path

        )

        self.workbook.close()