from openpyxl import load_workbook


# ==========================================================
# AI Curriculum Workbook
# Single Source of Truth
# ==========================================================

WORKBOOK_PATH = (
    "/volume1/docker/curriculum-builder/templates/"
    "AI_Curriculum_Workbook_v4.0_Production.xlsx"
)


class WorkbookReader:

    def __init__(self):
        self.wb = load_workbook(
            WORKBOOK_PATH,
            data_only=True
        )

    # ------------------------------------------------------

    def get_sheet(self, sheet_name):

        return self.wb[sheet_name]

    # ------------------------------------------------------

    def get_headers(self, sheet_name):

        ws = self.get_sheet(sheet_name)

        headers = {}

        for i, cell in enumerate(ws[1], start=1):
            headers[cell.value] = i

        return headers

    # ------------------------------------------------------

    def get_year_levels(self):

        ws = self.get_sheet("Curriculum_Master")

        headers = self.get_headers("Curriculum_Master")

        year_col = headers["Year_Level"]

        years = set()

        for row in range(2, ws.max_row + 1):

            value = ws.cell(row=row, column=year_col).value

            if value:
                years.add(str(value))

        return sorted(years)

    # ------------------------------------------------------

    def reload(self):

        self.wb = load_workbook(
            WORKBOOK_PATH,
            data_only=True
        )