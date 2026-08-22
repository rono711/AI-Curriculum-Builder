import base64
import html
import json
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(
        0,
        str(PROJECT_ROOT)
    )

from workbook_reader import WorkbookReader
from moodle_client import MoodleClient
from sync_writer import SyncWriter


# ==========================================================
# Publisher Builder
# ==========================================================

class PublisherBuilder:

    def __init__(self):

        self.reader = WorkbookReader()

        self.moodle = MoodleClient()

        self.sync = SyncWriter()

    # ======================================================
    # Helpers
    # ======================================================

    @staticmethod
    def _text(value):

        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def _read_text_file(path):

        path = Path(path)

        if not path.exists():
            return ""

        return path.read_text(
            encoding="utf-8"
        )

    @staticmethod
    def _read_json_field(
            path,
            field
    ):

        path = Path(path)

        if not path.exists():
            return ""

        try:

            data = json.loads(
                path.read_text(
                    encoding="utf-8"
                )
            )

        except (
            OSError,
            UnicodeDecodeError,
            json.JSONDecodeError
        ):
            return ""

        value = data.get(
            field,
            ""
        )

        if value is None:
            return ""

        return str(value).strip()

    # ======================================================
    # Lesson Content
    # ======================================================

    def _build_lesson_content(
            self,
            build_root,
            build_name
    ):

        lesson_file = (
            Path(build_root)
            / "Content"
            / build_name
            / "lesson_output.md"
        )

        lesson_markdown = self._read_text_file(
            lesson_file
        )

        if not lesson_markdown:

            lesson_json = (
                Path(build_root)
                / "Content"
                / build_name
                / "lesson_output.json"
            )

            lesson_markdown = self._read_json_field(
                lesson_json,
                "markdown"
            )

        if not lesson_markdown:
            return ""


        #
        # Preserve the existing Publisher behaviour for the
        # first production integration test.
        #

        return (
                "<pre>\n"
                + html.escape(lesson_markdown)
                + "\n</pre>"
        )

    # ======================================================
    # Did You Know / Gamma
    # ======================================================

    def _build_did_you_know(
            self,
            build_root,
            build_name,
            slides
    ):

        slides_folder = (
            Path(build_root)
            / "Slides"
            / build_name
        )

        urls_file = (
            slides_folder
            / "slides_urls.json"
        )

        embed_url = ""

        #
        # Prefer the generated Gamma URL file.
        #
        if urls_file.exists():

            data = json.loads(
                urls_file.read_text(
                    encoding="utf-8"
                )
            )

            embed_url = self._text(
                data.get("gamma_embed_url")
            )

        #
        # Workbook URL fallback.
        #
        if not embed_url:

            embed_url = self._text(
                slides.get("gamma_embed_url")
            )

        #
        # Complete embed HTML fallback.
        #
        if not embed_url:

            embed_html = self._text(
                slides.get("slides_embed_html")
            )

            if embed_html:
                return embed_html

        if not embed_url:
            return ""

        safe_url = html.escape(
            embed_url,
            quote=True
        )

        return (
            '<iframe '
            f'src="{safe_url}" '
            'width="100%" '
            'height="720" '
            'referrerpolicy="no-referrer" '
            'allow="fullscreen" '
            'allowfullscreen>'
            '</iframe>'
        )

    # ======================================================
    # Activities
    # ======================================================

    def _build_activities(
            self,
            build_root,
            build_name,
            activities
    ):

        activities_file = (
            Path(build_root)
            / "Activities"
            / build_name
            / "activities.html"
        )

        content = self._read_text_file(
            activities_file
        )

        if content:
            return content

        activities_json = (
            Path(build_root)
            / "Activities"
            / build_name
            / "activities.json"
        )

        content = self._read_json_field(
            activities_json,
            "html"
        )

        if content:
            return content

        workbook_value = self._text(
            activities.get("activities_html")
        )

        if workbook_value.lower().endswith(
                ".html"
        ):
            return ""

        return workbook_value

    # ======================================================
    # Recap
    # ======================================================

    def _build_recap(
            self,
            build_root,
            build_name,
            recap
    ):

        recap_file = (
            Path(build_root)
            / "Recap"
            / build_name
            / "recap.html"
        )

        content = self._read_text_file(
            recap_file
        )

        if content:
            return content

        recap_json = (
            Path(build_root)
            / "Recap"
            / build_name
            / "recap.json"
        )

        content = self._read_json_field(
            recap_json,
            "html"
        )

        if content:
            return content

        workbook_value = self._text(
            recap.get("recap_html")
        )

        if workbook_value.lower().endswith(
                ".html"
        ):
            return ""

        return workbook_value

    # ======================================================
    # Quiz Content
    # ======================================================

    def _build_quiz_content(
            self,
            build_root,
            build_name,
            quiz
    ):

        #
        # Prefer the GIFT content already stored in the workbook.
        #
        gift_content = self._text(
            quiz.get("gift_content")
        )

        if gift_content:
            return gift_content

        #
        # Fallback to generated .gift file.
        #
        gift_filename = self._text(
            quiz.get("gift_filename")
        )

        if not gift_filename:
            return ""

        gift_file = (
            Path(build_root)
            / "Quiz"
            / build_name
            / gift_filename
        )

        gift_content = self._read_text_file(
            gift_file
        )

        if gift_content:
            return gift_content

        gift_json = (
            Path(build_root)
            / "Quiz"
            / build_name
            / "lesson_quiz.json"
        )

        return self._read_json_field(
            gift_json,
            "gift"
        )

    # ======================================================
    # Moodle Course Mapping
    # ======================================================

    def _resolve_course(
            self,
            subject,
            year_level,
            build_id,
            lesson_package_id,
            school_level
    ):

        from openpyxl import load_workbook

        mapping_file = (
            PROJECT_ROOT
            / "data"
            / "moodle_course_mapping.xlsx"
        )

        if not mapping_file.exists():
            raise RuntimeError(
                "Moodle course mapping file not found: "
                f"{mapping_file}"
            )

        # ==================================================
        # First try existing mapping
        # ==================================================

        workbook = load_workbook(
            mapping_file,
            read_only=True,
            data_only=True
        )

        try:

            if "Course_Mapping" not in workbook.sheetnames:
                raise RuntimeError(
                    "Course_Mapping worksheet not found in "
                    "moodle_course_mapping.xlsx"
                )

            sheet = workbook["Course_Mapping"]

            headers = {}

            for cell in sheet[1]:

                if not cell.value:
                    continue

                key = (
                    str(cell.value)
                    .strip()
                    .lower()
                    .replace(" ", "_")
                    .replace("-", "_")
                )

                headers[key] = cell.column

            required_headers = [
                "subject",
                "year",
                "moodle_course_id",
            ]

            for field in required_headers:

                if field not in headers:
                    raise RuntimeError(
                        "Missing Moodle course mapping column: "
                        f"{field}"
                    )

            wanted_subject = self._text(
                subject
            ).casefold()

            wanted_year = self._text(
                year_level
            ).casefold()

            matches = []

            for row in range(
                    2,
                    sheet.max_row + 1
            ):

                mapped_subject = self._text(
                    sheet.cell(
                        row=row,
                        column=headers["subject"]
                    ).value
                )

                mapped_year = self._text(
                    sheet.cell(
                        row=row,
                        column=headers["year"]
                    ).value
                )

                if (
                    mapped_subject.casefold()
                    == wanted_subject
                    and
                    mapped_year.casefold()
                    == wanted_year
                ):

                    courseid = sheet.cell(
                        row=row,
                        column=headers[
                            "moodle_course_id"
                        ]
                    ).value

                    course_name = ""

                    if "moodle_course_name" in headers:

                        course_name = self._text(
                            sheet.cell(
                                row=row,
                                column=headers[
                                    "moodle_course_name"
                                ]
                            ).value
                        )

                    matches.append({
                        "courseid": int(courseid),
                        "name": course_name,
                    })

            if len(matches) > 1:

                raise RuntimeError(
                    "Multiple Moodle course mappings found for "
                    f"subject={subject!r}, "
                    f"year_level={year_level!r}"
                )

            if matches:

                print("=" * 60)
                print("EXISTING MOODLE COURSE MAPPING FOUND")
                print("Subject :", subject)
                print("Year    :", year_level)
                print("Course  :", matches[0]["courseid"])
                print("=" * 60)

                return matches[0]

        finally:

            workbook.close()

        # ==================================================
        # Mapping missing - create/reuse Moodle course
        # ==================================================

        print("=" * 60)
        print("MOODLE COURSE MAPPING NOT FOUND")
        print("Creating or reusing Moodle course")
        print("Subject :", subject)
        print("Year    :", year_level)
        print("=" * 60)

        result = self.moodle.ensure_course(
            self._text(school_level),
            self._text(subject),
            self._text(year_level)
        )

        if not isinstance(result, dict):

            raise RuntimeError(
                "Moodle publish_course returned "
                "an invalid response."
            )

        if result.get("status") != "SUCCESS":

            raise RuntimeError(
                "Moodle course creation/reuse failed: "
                f"{result}"
            )

        courseid = result.get("courseid")

        if not courseid:

            raise RuntimeError(
                "Moodle publish_course did not return "
                "a courseid."
            )

        course_name = self._text(
            result.get("fullname")
        )

        # ==================================================
        # Save new Moodle mapping
        # ==================================================

        workbook = load_workbook(
            mapping_file
        )

        try:

            sheet = workbook["Course_Mapping"]

            sheet.append([
                self._text(subject),
                self._text(year_level),
                int(courseid),
                course_name,
            ])

            workbook.save(
                mapping_file
            )

        finally:

            workbook.close()

        print("=" * 60)
        print("NEW MOODLE COURSE MAPPING SAVED")
        print("Subject :", subject)
        print("Year    :", year_level)
        print("Course  :", courseid)
        print("Name    :", course_name)
        print("=" * 60)

        return {
            "courseid": int(courseid),
            "name": course_name,
    
        }
    # Publish
    # ======================================================

    def publish(
            self,
            build_root,
            build_name,
            lesson_package_id
    ):

        workbook = (
            Path(build_root)
            / "Workbook"
            / f"{build_name}.xlsx"
        )

        print("=" * 60)
        print("PUBLISHER WORKBOOK")
        print(workbook)
        print("=" * 60)

        lesson = self.reader.read(
            workbook,
            lesson_package_id
        )

        metadata = lesson["metadata"]
        descriptions = lesson["descriptions"]
        slides = lesson["slides"]
        quiz = lesson["quiz"]
        activities = lesson["activities"]
        recap = lesson["recap"]

        print("=" * 60)
        print("LESSON METADATA")
        print(metadata)
        print("=" * 60)

        # ==================================================
        # Resolve existing Moodle course
        # ==================================================

        course = self._resolve_course(
            metadata["subject"],
            metadata["year_level"],
            build_name,
            lesson_package_id,
            metadata.get("school_level", "")
        )

        courseid = int(
            course["courseid"]
        )

        print("=" * 60)
        print("MOODLE COURSE MAPPING")
        print(
            "Subject:",
            metadata["subject"]
        )
        print(
            "Year:",
            metadata["year_level"]
        )
        print(
            "Course ID:",
            courseid
        )
        print(
            "Course Name:",
            course.get("name", "")
        )
        print("=" * 60)

        # ==================================================
        # Build generated lesson assets
        # ==================================================

        lesson_content = self._build_lesson_content(
            build_root,
            build_name
        )

        did_you_know = self._build_did_you_know(
            build_root,
            build_name,
            slides
        )

        quiz_content = self._build_quiz_content(
            build_root,
            build_name,
            quiz
        )

        activities_content = self._build_activities(
            build_root,
            build_name,
            activities
        )

        recap_content = self._build_recap(
            build_root,
            build_name,
            recap
        )

        # ==================================================
        # Validation
        # ==================================================

        if not lesson_content:
            raise RuntimeError(
                "Lesson Content is empty."
            )

        if not quiz_content:
            raise RuntimeError(
                "Quiz GIFT content is empty."
            )

        if not activities_content:
            raise RuntimeError(
                "Activities content is empty."
            )

        if not recap_content:
            raise RuntimeError(
                "Recap content is empty."
            )

        # ==================================================
        # Lesson title
        # ==================================================

        display_title = self._text(
            descriptions.get("display_title")
        )

        if not display_title:
            display_title = self._text(
                metadata.get("lesson_title")
            )

        if not display_title:
            display_title = self._text(
                metadata.get("elaboration")
            )

        curriculum_code = self._text(
            metadata.get("curriculum_code")
        )

        elaboration = self._text(
            metadata.get("elaboration")
        )

        if not curriculum_code:
            raise RuntimeError(
                "Curriculum code is empty."
            )

        if not elaboration:
            raise RuntimeError(
                "Curriculum elaboration is empty."
            )

        lesson_title = display_title

        # ==================================================
        # Complete Moodle lesson payload
        # ==================================================
        # ==================================================
        # Lesson Elaboration Image
        #
        # ONE image belongs to this curriculum elaboration.
        # ==================================================

        parent_code = self._text(
            metadata.get("parent_code")
        )

        if not parent_code:

            raise RuntimeError(
                "Parent curriculum code is empty."
            )

        elaboration_image = (
            Path(build_root)
            / "Images"
            / build_name
            / f"{curriculum_code}_elaboration.png"
        )

        if not elaboration_image.exists():

            raise RuntimeError(
                "Lesson elaboration image not found: "
                f"{elaboration_image}"
            )

        image_bytes = (
            elaboration_image.read_bytes()
        )

        if not image_bytes:

            raise RuntimeError(
                "Lesson elaboration image is empty: "
                f"{elaboration_image}"
            )

        elaboration_image_base64 = (
            base64.b64encode(
                image_bytes
            ).decode("ascii")
        )

        elaboration_image_name = (
            elaboration_image.name
        )

        
        print("=" * 60)
        print("LESSON ELABORATION IMAGE")
        print("Parent Code:", parent_code)
        print("Image:", elaboration_image)
        print("Bytes:", len(image_bytes))
        print("=" * 60)

        payload = {

            "courseid":
                courseid,

            "strand":
                self._text(
                    metadata.get("strand")
                ),

            "substrand":
                self._text(
                    metadata.get("sub_strand")
                ),

            "curriculumcode":
                curriculum_code,

            "elaboration":
                elaboration,

            "contentdescription":
                self._text(
                    metadata.get(
                        "content_description"
                    )
                ),

            "parentcode":
                parent_code,

            "contentdescriptionimagename":
                elaboration_image_name,

            "contentdescriptionimage":
                elaboration_image_base64,
                
            "lesson[title]":
                lesson_title,

            "lesson[lessoncontent]":
                lesson_content,

            "lesson[lessondescription]":
                self._text(
                    descriptions.get(
                        "mission_description"
                    )
                ),

            "lesson[didyouknow]":
                did_you_know,

            "lesson[didyouknowdescription]":
                self._text(
                    descriptions.get(
                        "slides_description"
                    )
                ),

            "lesson[quiztitle]":
                self._text(
                    quiz.get("quiz_title")
                )
                or
                self._text(
                    descriptions.get(
                        "quiz_title"
                    )
                )
                or
                "Checking Your Thinking",

            "lesson[quizdescription]":
                self._text(
                    descriptions.get(
                        "quiz_description"
                    )
                )
                or
                self._text(
                    quiz.get(
                        "quiz_description"
                    )
                ),

            "lesson[quizformat]":
                "gift",

            "lesson[quizcontent]":
                quiz_content,

            "lesson[activities]":
                activities_content,

            "lesson[activitiesdescription]":
                self._text(
                    descriptions.get(
                        "activities_description"
                    )
                )
                or
                self._text(
                    activities.get(
                        "activities_description"
                    )
                ),

            "lesson[recap]":
                recap_content,

            "lesson[recapdescription]":
                self._text(
                    descriptions.get(
                        "recap_description"
                    )
                )
                or
                self._text(
                    recap.get(
                        "recap_description"
                    )
                ),
        }

        # ==================================================
        # Validate structural metadata
        # ==================================================

        for required in [
            "strand",
            "substrand",
            "contentdescription",
            "lesson[title]",
        ]:

            if not payload[required]:

                raise RuntimeError(
                    f"Required Moodle field is empty: "
                    f"{required}"
                )

        print("=" * 60)
        print("RONO PUBLISHER PAYLOAD")

        safe_payload = {
            **payload,

            "contentdescriptionimage":
                (
                    f"<BASE64 "
                    f"{len(elaboration_image_base64)} chars>"
                ),

            "lesson[quizcontent]":
                f"<GIFT {len(quiz_content)} chars>",

            "lesson[lessoncontent]":
                f"<CONTENT {len(lesson_content)} chars>",

            "lesson[activities]":
                f"<ACTIVITIES {len(activities_content)} chars>",

            "lesson[recap]":
                f"<RECAP {len(recap_content)} chars>",
        }

        print(safe_payload)
        print("=" * 60)

        # ==================================================
        # ONE atomic Moodle lesson publication
        # ==================================================

        published = self.moodle.publish_lesson(
            payload
        )

        print("=" * 60)
        print("RONO PUBLISHER RESULT")
        print(published)
        print("=" * 60)

        # ==================================================
        # Verify Moodle result
        # ==================================================

        if (
            published.get("status")
            != "success"
        ):

            raise RuntimeError(
                "Moodle lesson publishing did not "
                "return success."
            )

        if int(
            published.get(
                "questioncount",
                0
            )
        ) <= 0:

            raise RuntimeError(
                "Moodle published the lesson but "
                "reported zero imported questions."
            )

        if int(
            published.get(
                "slotcount",
                0
            )
        ) <= 0:

            raise RuntimeError(
                "Moodle published the Quiz but "
                "reported zero Quiz slots."
            )

        # ==================================================
        # Update Moodle_Publish worksheet
        # ==================================================

        self.sync.update(
            workbook,
            lesson_package_id,
            published
        )

        print("=" * 60)
        print("MOODLE_PUBLISH WORKBOOK UPDATED")
        print(lesson_package_id)
        print("=" * 60)

        # ==================================================
        # Return
        # ==================================================

        return {

            "status":
                "SUCCESS",

            "lesson_package_id":
                lesson_package_id,

            "course":
                course,

            "publisher":
                published,

            "quiz": {

                "quizid":
                    published.get(
                        "quizid"
                    ),

                "cmid":
                    published.get(
                        "quizcmid"
                    ),

                "contextid":
                    published.get(
                        "quizcontextid"
                    ),

                "questioncount":
                    published.get(
                        "questioncount"
                    ),

                "slotcount":
                    published.get(
                        "slotcount"
                    ),

                "sumgrades":
                    published.get(
                        "quizsumgrades"
                    ),
            },

            "mission": {

                "cmid":
                    published.get(
                        "lessoncontentcmid"
                    )
            },

            "did_you_know": {

                "cmid":
                    published.get(
                        "didyouknowcmid"
                    )
            },

            "activities": {

                "cmid":
                    published.get(
                        "activitiescmid"
                    )
            },

            "recap": {

                "cmid":
                    published.get(
                        "recapcmid"
                    )
            },
        }
            # ======================================================
    # Publish Selective UPDATE
    # ======================================================

    def publish_update(
            self,
            build_root,
            build_name,
            lesson_package_id,
            update_components,
            moodle_identity
    ):

        """
        Update selected components of an existing published
        Moodle lesson without creating new Moodle activities.

        Initial supported component:

            recap
        """

        update_components = (
            update_components
            or []
        )

        update_components = [
            str(component).strip().lower()
            for component in update_components
            if str(component).strip()
        ]

        if not update_components:

            raise ValueError(
                "UPDATE requires at least one component."
            )

        supported_components = {
            "lesson_content",
            "slides",
            "activities",
            "recap",
            "image",
        }

        unsupported = (
            set(update_components)
            - supported_components
        )

        if unsupported:

            raise ValueError(
                "Publisher UPDATE does not support: "
                + ", ".join(
                    sorted(unsupported)
                )
            )

        if not moodle_identity:

            raise ValueError(
                "Moodle identity is required for UPDATE."
            )

        courseid = moodle_identity.get(
            "moodle_course_id"
        )

        if not courseid:

            raise ValueError(
                "UPDATE Moodle course ID is missing."
            )

        results = {}
        


        lesson = None
        descriptions = {}

        if any(
            component in update_components
            for component in (
                "lesson_content",
                "slides",
                "activities",
                "recap",
                "image",
            )
        ):

            workbook = (
                Path(build_root)
                / "Workbook"
                / f"{build_name}.xlsx"
            )

            lesson = self.reader.read(
                workbook,
                lesson_package_id
            )

            descriptions = (
                lesson.get("descriptions")
                or {}
            )

        # ==================================================
        # Elaboration Text & Media Banner / Image
        # ==================================================

        if "image" in update_components:

            banner_cmid = (
                moodle_identity.get(
                    "moodle_content_description_cmid"
                )
            )

            if not banner_cmid:
                raise ValueError(
                    "UPDATE elaboration banner CMID is missing."
                )

            metadata = (
                lesson.get("metadata")
                or {}
            )

            curriculum_code = self._text(
                metadata.get("curriculum_code")
            )

            parent_code = self._text(
                metadata.get("parent_code")
            )

            content_description = self._text(
                metadata.get("content_description")
            )

            elaboration = self._text(
                metadata.get("elaboration")
            )

            if not curriculum_code:
                raise RuntimeError(
                    "UPDATE curriculum code is empty."
                )

            if not parent_code:
                raise RuntimeError(
                    "UPDATE parent code is empty."
                )

            if not content_description:
                raise RuntimeError(
                    "UPDATE content description is empty."
                )

            if not elaboration:
                raise RuntimeError(
                    "UPDATE curriculum elaboration is empty."
                )

            elaboration_image = (
                Path(build_root)
                / "Images"
                / build_name
                / f"{curriculum_code}_elaboration.png"
            )

            if not elaboration_image.exists():
                raise RuntimeError(
                    "UPDATE elaboration image not found: "
                    f"{elaboration_image}"
                )

            image_bytes = (
                elaboration_image.read_bytes()
            )

            if not image_bytes:
                raise RuntimeError(
                    "UPDATE elaboration image is empty: "
                    f"{elaboration_image}"
                )

            elaboration_image_base64 = (
                base64.b64encode(
                    image_bytes
                ).decode("ascii")
            )

            print("=" * 60)
            print("PUBLISHER SELECTIVE UPDATE")
            print("Component : image")
            print("Course    :", courseid)
            print("CMID      :", banner_cmid)
            print("Code      :", curriculum_code)
            print("Parent    :", parent_code)
            print("Image     :", elaboration_image)
            print("Bytes     :", len(image_bytes))
            print("=" * 60)

            result = (
                self.moodle.update_elaboration_banner({
                    "courseid":
                        int(courseid),

                    "cmid":
                        int(banner_cmid),

                    "curriculumcode":
                        curriculum_code,

                    "parentcode":
                        parent_code,

                    "contentdescription":
                        content_description,

                    "elaboration":
                        elaboration,

                    "imagename":
                        elaboration_image.name,

                    "image":
                        elaboration_image_base64,
                })
            )

            if result.get("status") != "success":
                raise RuntimeError(
                    "Moodle elaboration banner UPDATE "
                    "did not return success: "
                    f"{result}"
                )

            if (
                int(result.get("cmid", 0))
                !=
                int(banner_cmid)
            ):
                raise RuntimeError(
                    "Moodle elaboration banner UPDATE "
                    "returned an unexpected CMID."
                )

            results["image"] = result

        # ==================================================
        # Lesson Content / Mission
        # ==================================================

        if "lesson_content" in update_components:

            lesson_content_cmid = (
                moodle_identity.get(
                    "moodle_lesson_content_cmid"
                )
            )

            if not lesson_content_cmid:

                raise ValueError(
                    "UPDATE Lesson Content CMID is missing."
                )

            lesson_content = (
                self._build_lesson_content(
                    build_root,
                    build_name
                )
            )

            if not lesson_content:

                raise RuntimeError(
                    "Generated Lesson Content is empty."
                )

            mission_description = (
                descriptions.get(
                    "mission_description"
                )
                or ""
            ).strip()

            if not mission_description:

                raise RuntimeError(
                    "Generated Mission description is empty."
                )
            print("=" * 60)
            print("PUBLISHER SELECTIVE UPDATE")
            print("Component : lesson_content")
            print("Course    :", courseid)
            print(
                "CMID      :",
                lesson_content_cmid
            )
            print(
                "Content   :",
                len(lesson_content),
                "chars"
            )
            print("=" * 60)

            result = (
                self.moodle.update_component({
                    "courseid":
                        int(courseid),

                    "component":
                        "lesson_content",

                    "cmid":
                        int(
                            lesson_content_cmid
                        ),

                    "content":
                        lesson_content,

                    "description":
                        mission_description,
                })
            )

            if (
                result.get("status")
                != "success"
            ):

                raise RuntimeError(
                    "Moodle Lesson Content UPDATE "
                    "did not return success: "
                    f"{result}"
                )

            if (
                int(
                    result.get(
                        "cmid",
                        0
                    )
                )
                !=
                int(lesson_content_cmid)
            ):

                raise RuntimeError(
                    "Moodle Lesson Content UPDATE "
                    "returned an unexpected CMID."
                )

            results["lesson_content"] = (
                result
            )

        # ==================================================
        # Slides / Did You Know
        # ==================================================

        if "slides" in update_components:

            did_you_know_cmid = (
                moodle_identity.get(
                    "moodle_did_you_know_cmid"
                )
            )

            if not did_you_know_cmid:

                raise ValueError(
                    "UPDATE Did You Know CMID is missing."
                )

            did_you_know_content = (
                self._build_did_you_know(
                    build_root,
                    build_name,
                    lesson["slides"]
                )
            )

            if not did_you_know_content:

                raise RuntimeError(
                    "Generated Did You Know content is empty."
                )
            slides_description = (
                descriptions.get(
                    "slides_description"
                )
                or ""
            ).strip()

            if not slides_description:

                raise RuntimeError(
                    "Generated Slides description is empty."
                )

            print("=" * 60)
            print("PUBLISHER SELECTIVE UPDATE")
            print("Component : slides")
            print("Course    :", courseid)
            print(
                "CMID      :",
                did_you_know_cmid
            )
            print(
                "Content   :",
                len(did_you_know_content),
                "chars"
            )
            print("=" * 60)

            result = (
                self.moodle.update_component({
                    "courseid":
                        int(courseid),

                    "component":
                        "slides",

                    "cmid":
                        int(did_you_know_cmid),

                    "content":
                        did_you_know_content,

                    "description":
                        slides_description,
                })
            )

            if result.get("status") != "success":

                raise RuntimeError(
                    "Moodle Did You Know UPDATE "
                    "did not return success: "
                    f"{result}"
                )

            if (
                int(result.get("cmid", 0))
                !=
                int(did_you_know_cmid)
            ):

                raise RuntimeError(
                    "Moodle Did You Know UPDATE "
                    "returned an unexpected CMID."
                )

            results["slides"] = result

        # ==================================================
        # Activities / Let's Do It
        # ==================================================

        if "activities" in update_components:

            activities_cmid = (
                moodle_identity.get(
                    "moodle_activities_cmid"
                )
            )

            if not activities_cmid:

                raise ValueError(
                    "UPDATE Activities CMID is missing."
                )

            activities_content = (
                self._build_activities(
                    build_root,
                    build_name,
                    lesson["activities"]
                )
            )
            
            if not activities_content:

                raise RuntimeError(
                    "Generated Activities content is empty."
                )

            activities_description = (
                descriptions.get(
                    "activities_description"
                )
                or ""
            ).strip()

            if not activities_description:

                raise RuntimeError(
                    "Generated Activities description is empty."
                )

            print("=" * 60)
            print("PUBLISHER SELECTIVE UPDATE")
            print("Component : activities")
            print("Course    :", courseid)
            print(
                "CMID      :",
                activities_cmid
            )
            print(
                "Content   :",
                len(activities_content),
                "chars"
            )
            print("=" * 60)

            result = (
                self.moodle.update_component({
                    "courseid":
                        int(courseid),

                    "component":
                        "activities",

                    "cmid":
                        int(activities_cmid),

                    "content":
                        activities_content,
                     "description":
                        activities_description,
                    })
            )

            if result.get("status") != "success":

                raise RuntimeError(
                    "Moodle Activities UPDATE "
                    "did not return success: "
                    f"{result}"
                )

            if (
                int(result.get("cmid", 0))
                !=
                int(activities_cmid)
            ):

                raise RuntimeError(
                    "Moodle Activities UPDATE "
                    "returned an unexpected CMID."
                )

            results["activities"] = result
       
       # ==================================================
        # Recap
        # ==================================================

        if "recap" in update_components:

            recap_cmid = moodle_identity.get(
                "moodle_recap_cmid"
            )

            if not recap_cmid:

                raise ValueError(
                    "UPDATE Recap CMID is missing."
                )

            recap_file = (
                Path(build_root)
                / "Recap"
                / build_name
                / "recap.html"
            )

            recap_content = (
                self._read_text_file(
                    recap_file
                )
            )

            if not recap_content:

                raise RuntimeError(
                    "Generated Recap HTML is empty: "
                    + str(recap_file)
                )

            recap_description = (
                descriptions.get(
                    "recap_description"
                )
                or ""
            ).strip()

            if not recap_description:

                raise RuntimeError(
                    "Generated Recap description is empty."
                )

            print("=" * 60)
            print("PUBLISHER SELECTIVE UPDATE")
            print("Component : recap")
            print("Course    :", courseid)
            print("CMID      :", recap_cmid)
            print(
                "Content   :",
                len(recap_content),
                "chars"
            )
            print("=" * 60)

            result = (
                self.moodle.update_component({
                    "courseid":
                        int(courseid),

                    "component":
                        "recap",

                    "cmid":
                        int(recap_cmid),

                    "content":
                        recap_content,
                    "description":
                        recap_description,
                    })
            )

            if (
                result.get("status")
                != "success"
            ):

                raise RuntimeError(
                    "Moodle Recap UPDATE did not "
                    "return success: "
                    f"{result}"
                )

            if (
                int(
                    result.get(
                        "cmid",
                        0
                    )
                )
                !=
                int(recap_cmid)
            ):

                raise RuntimeError(
                    "Moodle Recap UPDATE returned "
                    "an unexpected CMID."
                )

            results["recap"] = result

        # ==================================================
        # Complete
        # ==================================================

        print("=" * 60)
        print("PUBLISHER UPDATE COMPLETE")
        print(
            "Lesson:",
            lesson_package_id
        )
        print(
            "Components:",
            update_components
        )
        print("=" * 60)

        return {
            "status":
                "SUCCESS",

            "build_mode":
                "UPDATE",

            "lesson_package_id":
                lesson_package_id,

            "update_components":
                update_components,

            "courseid":
                int(courseid),

            "components":
                results,
        }
