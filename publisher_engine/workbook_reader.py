from openpyxl import load_workbook


# ==========================================================
# Workbook Reader
# ==========================================================

class WorkbookReader:

    # ======================================================
    # Header Map
    # ======================================================

    @staticmethod
    def header_map(sheet):

        headers = {}

        for cell in sheet[1]:

            if not cell.value:
                continue

            key = (
                str(cell.value)
                .strip()
                .lower()
                .replace(" ", "_")
                .replace("-", "_")
            )

            headers[key] = cell.column

        return headers

    # ======================================================
    # Read One Worksheet
    # ======================================================

    def read_sheet(

            self,

            workbook,

            worksheet,

            lesson_package_id

    ):

        if worksheet not in workbook.sheetnames:

            return {}

        sheet = workbook[worksheet]

        headers = self.header_map(

            sheet

        )

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:

                break

            if str(value) == str(

                    lesson_package_id

            ):

                record = {}

                for key, column in headers.items():
                    record[key] = sheet.cell(
                        row=row,
                        column=column
                    ).value

                return record

            row += 1

        return {}

    # ======================================================
    # Read Lesson
    # ======================================================

    # ======================================================
    # Read Lesson
    # ======================================================

    def read(

            self,

            workbook_path,

            lesson_package_id

    ):

        workbook = load_workbook(

            workbook_path,

            data_only=True

        )

        metadata = self.read_sheet(

            workbook,

            "Lesson_Metadata",

            lesson_package_id

        )

        # --------------------------------------------------
        # Parent Content Description identity
        #
        # parent_code belongs to Lesson_DB rather than
        # Lesson_Metadata. Resolve it using this exact
        # lesson_package_id only.
        # --------------------------------------------------

        lesson_db = self.read_sheet(

            workbook,

            "Lesson_DB",

            lesson_package_id

        )

        parent_code = (
            lesson_db.get(
                "parent_code",
                ""
            )
            or ""
        )

        metadata["parent_code"] = (
            str(parent_code).strip()
        )
        #
        # --------------------------------------------------
        # Derive School Level
        # --------------------------------------------------
        #

        year = str(

            metadata.get(

                "year_level",

                ""

            )

        ).strip()

        if year in [
            "Foundation",
            "Foundation Year",
            "Year 1",
            "Year 2"
        ]:

            metadata["school_level"] = "Lower Primary"

        elif year in [
            "Year 3",
            "Year 4",
            "Year 5",
            "Year 6"
        ]:

            metadata["school_level"] = "Upper Primary"

        elif year in [
            "Year 7",
            "Year 8",
            "Year 9",
            "Year 10"
        ]:

            metadata["school_level"] = "Secondary"

        elif year in [
            "Year 11",
            "Year 12"
        ]:

            metadata["school_level"] = "Senior Secondary"

        else:

            raise ValueError(
                f"Unsupported Year Level: {year!r}"
            )
        lesson = {

            "metadata":

                metadata,

            "slides":

                self.read_sheet(

                    workbook,

                    "Gamma_Slides",

                    lesson_package_id

                ),

            "quiz":

                self.read_sheet(

                    workbook,

                    "Quiz",

                    lesson_package_id

                ),

            "activities":

                self.read_sheet(

                    workbook,

                    "Activities",

                    lesson_package_id

                ),

            "recap":

                self.read_sheet(

                    workbook,

                    "Recap",

                    lesson_package_id

                ),

            "descriptions":

                self.read_sheet(

                    workbook,

                    "Descriptions",

                    lesson_package_id

                ),

            "publish":

                self.read_sheet(

                    workbook,

                    "Moodle_Publish",

                    lesson_package_id

                )

        }

        workbook.close()

        return lesson
