from datetime import datetime

from openpyxl import load_workbook


# ==========================================================
# Sync Writer
# ==========================================================

class SyncWriter:

    # ======================================================
    # Header Map
    # ======================================================

    @staticmethod
    def header_map(sheet):

        headers = {}

        for cell in sheet[1]:

            if cell.value:

                headers[
                    str(cell.value).strip()
                ] = cell.column

        return headers

    # ======================================================
    # Update Workbook
    # ======================================================

    def update(

            self,

            workbook_path,

            lesson_package_id,

            published

    ):

        """
        Update Moodle_Publish using the result returned by:

            local_rono_publisher_publish_lesson

        The Moodle plugin now publishes the complete lesson
        atomically, so separate section/page/quiz result
        dictionaries are no longer required.
        """

        workbook = load_workbook(
            workbook_path
        )

        sheet = workbook[
            "Moodle_Publish"
        ]

        headers = self.header_map(
            sheet
        )

        row = 2

        found = False

        while True:

            value = sheet.cell(
                row=row,
                column=headers[
                    "lesson_package_id"
                ]
            ).value

            if not value:
                break

            if str(value) == str(
                    lesson_package_id
            ):

                found = True

                # ==========================================
                # Course
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "course_id",
                    published.get(
                        "courseid",
                        ""
                    )
                )

                # ==========================================
                # Strand Section
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "section_id",
                    published.get(
                        "strandsectionid",
                        ""
                    )
                )

                # ==========================================
                # Mission / Lesson Content
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "mission_cmid",
                    published.get(
                        "lessoncontentcmid",
                        ""
                    )
                )

                # ==========================================
                # Quiz
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "quizid",
                    published.get(
                        "quizid",
                        ""
                    )
                )

                self.write(
                    sheet,
                    headers,
                    row,
                    "quiz_cmid",
                    published.get(
                        "quizcmid",
                        ""
                    )
                )

                # ==========================================
                # Activities
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "activities_cmid",
                    published.get(
                        "activitiescmid",
                        ""
                    )
                )

                # ==========================================
                # Recap
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "recap_cmid",
                    published.get(
                        "recapcmid",
                        ""
                    )
                )

                # ==========================================
                # Publication status
                # ==========================================

                self.write(
                    sheet,
                    headers,
                    row,
                    "publication_status",
                    "PUBLISHED"
                )

                self.write(
                    sheet,
                    headers,
                    row,
                    "last_synced",
                    datetime.now()
                )

                self.write(
                    sheet,
                    headers,
                    row,
                    "needs_sync",
                    "NO"
                )

                break

            row += 1

        if not found:

            workbook.close()

            raise RuntimeError(
                "Lesson package was not found in "
                "Moodle_Publish worksheet: "
                f"{lesson_package_id}"
            )

        workbook.save(
            workbook_path
        )

        workbook.close()

    # ======================================================
    # Write
    # ======================================================

    def write(

            self,

            sheet,

            headers,

            row,

            field,

            value

    ):

        #
        # Some older workbook templates may not contain every
        # Moodle publishing field. Missing optional fields are
        # deliberately ignored.
        #

        if field in headers:

            sheet.cell(
                row=row,
                column=headers[field]
            ).value = value