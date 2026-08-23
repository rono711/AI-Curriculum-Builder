from pathlib import Path
import sys

import requests
from openpyxl import load_workbook



# ==========================================================
# Shared Project Modules
# ==========================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


from build_registry import (
    start_build,
    mark_published,
    mark_update_ready,
    get_published_build,
    mark_generated,
    mark_failed,
    get_previous_published_build,
    inherit_moodle_identity,
    register_quiz_questions
)

from config import (
    PROMPT_ENGINE_URL,
    GAMMA_ENGINE_URL,
    QUIZ_ENGINE_URL,
    ACTIVITIES_ENGINE_URL,
    RECAP_ENGINE_URL,
    IMAGE_ENGINE_URL,
    PUBLISHER_ENGINE_URL,
    PROMPT_TIMEOUT,
    ENGINE_TIMEOUT
)


# ==========================================================
# Pipeline Builder
# ==========================================================

class PipelineBuilder:

    def __init__(self):
        print("=" * 60)
        print("PIPELINE BUILDER INITIALIZED")
        print("=" * 60)
    # ======================================================
    # Moodle Publication Status
    # ======================================================

    @staticmethod
    def _get_publication_status(
            workbook_path,
            lesson_package_id
    ):

        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=True
        )

        try:

            if "Moodle_Publish" not in workbook.sheetnames:
                return ""

            sheet = workbook[
                "Moodle_Publish"
            ]

            headers = {}

            for cell in sheet[1]:

                if not cell.value:
                    continue

                key = (
                    str(cell.value)
                    .strip()
                    .lower()
                    .replace(" ", "_")
                    .replace("-", "_")
                )

                headers[key] = cell.column

            if "lesson_package_id" not in headers:
                return ""

            if "publication_status" not in headers:
                return ""

            for row in range(
                    2,
                    sheet.max_row + 1
            ):

                value = sheet.cell(
                    row=row,
                    column=headers[
                        "lesson_package_id"
                    ]
                ).value

                if str(value) == str(
                        lesson_package_id
                ):

                    status = sheet.cell(
                        row=row,
                        column=headers[
                            "publication_status"
                        ]
                    ).value

                    if status is None:
                        return ""

                    return str(
                        status
                    ).strip().upper()

            return ""

        finally:

            workbook.close()
    
    # ======================================================
    # Moodle Publication Status Writer
    # ======================================================

    @staticmethod
    def _set_publication_status(
            workbook_path,
            lesson_package_id,
            status,
            needs_sync="YES"
    ):

        workbook = load_workbook(
            workbook_path
        )

        try:

            if "Moodle_Publish" not in workbook.sheetnames:
                raise RuntimeError(
                    "Moodle_Publish worksheet not found."
                )

            sheet = workbook[
                "Moodle_Publish"
            ]

            headers = {}

            for cell in sheet[1]:

                if not cell.value:
                    continue

                key = (
                    str(cell.value)
                    .strip()
                    .lower()
                    .replace(" ", "_")
                    .replace("-", "_")
                )

                headers[key] = cell.column

            required = [
                "lesson_package_id",
                "publication_status",
                "needs_sync"
            ]

            for column in required:

                if column not in headers:
                    raise RuntimeError(
                        f"Moodle_Publish column missing: {column}"
                    )

            found = False

            for row in range(
                    2,
                    sheet.max_row + 1
            ):

                value = sheet.cell(
                    row=row,
                    column=headers[
                        "lesson_package_id"
                    ]
                ).value

                if str(value) != str(
                        lesson_package_id
                ):
                    continue

                sheet.cell(
                    row=row,
                    column=headers[
                        "publication_status"
                    ]
                ).value = status

                sheet.cell(
                    row=row,
                    column=headers[
                        "needs_sync"
                    ]
                ).value = needs_sync

                found = True
                break

            if not found:
                raise RuntimeError(
                    "Lesson package not found in "
                    "Moodle_Publish: "
                    f"{lesson_package_id}"
                )

            workbook.save(
                workbook_path
            )

        finally:

            workbook.close()


    # ======================================================
    # Safe Moodle Publisher Call
    # ======================================================

    @staticmethod
    def _call_publisher_safely(
            engine_build_root,
            engine_build_name,
            lesson_package_id
    ):

        try:

            response = requests.post(
                PUBLISHER_ENGINE_URL,
                json={
                    "build_root":
                        engine_build_root,

                    "build_name":
                        engine_build_name,

                    "lesson_package_id":
                        lesson_package_id
                },
                timeout=ENGINE_TIMEOUT
            )

            print(
                "Publisher",
                response.status_code
            )

            if response.status_code != 200:
                print(response.text)

            response.raise_for_status()

            publisher_result = response.json()

            if (
                publisher_result.get("status")
                != "SUCCESS"
            ):

                raise RuntimeError(
                    "Publisher Engine did not "
                    "return SUCCESS for "
                    f"{lesson_package_id}: "
                    f"{publisher_result}"
                )

            return {
                "success": True,
                "result": publisher_result
            }

        except Exception as exc:

            print("=" * 60)
            print("MOODLE PUBLICATION FAILED")
            print("Lesson :", lesson_package_id)
            print("Error  :", str(exc))
            print(
                "Content generation remains successful."
            )
            print("=" * 60)

            return {
                "success": False,
                "result": {
                    "status": "FAILED",
                    "reason":
                        "MOODLE_PUBLICATION_FAILED",
                    "error":
                        str(exc)
                }
            }

    # ======================================================
    # Build Progress Reporter
    # ======================================================

    @staticmethod
    def _report_progress(
            progress_url,
            progress_job_id,
            stage,
            message,
            percent
    ):

        if not progress_url:
            return

        if not progress_job_id:
            return

        callback_url = (
            progress_url.rstrip("/")
            + "/api/build-progress/"
            + str(progress_job_id)
        )

        try:

            response = requests.post(
                callback_url,
                json={
                    "stage":
                        stage,

                    "message":
                        message,

                    "percent":
                        int(percent),
                },
                timeout=5
            )

            print(
                "PROGRESS CALLBACK:",
                percent,
                stage,
                response.status_code
            )

        except Exception as exc:

            print(
                "PROGRESS CALLBACK WARNING:",
                str(exc)
            )


    # ======================================================
    # Run Pipeline
    # ======================================================

    def run(

            self,

            build_root,

            build_name,

            lesson_rows,

            progress_job_id="",

            progress_url=""

    ):

        self._report_progress(
            progress_url,
            progress_job_id,
            "PIPELINE_START",
            "Preparing lesson pipeline...",
            20
        )

        #
        # ==================================================
        # Prompt Generation Pipeline
        # ==================================================
        #
        build = {
            "path": Path(build_root) / "Workbook" / f"{build_name}.xlsx",
            "build_id": build_name,
            "filename": f"{build_name}.xlsx"
        }

        results = []

        for lesson in lesson_rows:

            lesson_package_id = lesson["lesson_package_id"]

            elaboration_key = lesson["elaboration_key"]

            print("=" * 60)
            print("PROCESSING", lesson_package_id)
            print("ELABORATION KEY:", elaboration_key)
            print("=" * 60)

            registry_record_id = start_build(

                elaboration_key=elaboration_key,

                learning_area=lesson["learning_area"],

                subject=lesson["subject"],

                year_level=lesson["year_level"],

                strand=lesson["strand"],

                sub_strand=lesson["sub_strand"],

                parent_code=lesson["parent_code"],

                topic_id=lesson["topic_id"],

                curriculum_code=lesson["curriculum_code"],

                content_description=lesson["content_description"],

                elaboration=lesson["elaboration"],

                build_id=lesson["build_id"],

                lesson_package_id=lesson_package_id,

                build_mode=lesson.get(
                    "build_mode",
                    "NEW"
                ),

                update_components=lesson.get(
                    "update_components",
                    []
                )

            )

            print(
                "REGISTRY RECORD:",
                registry_record_id,
                "STATUS: BUILDING"
            )

            # ==================================================
            # Build Mode / Selective UPDATE
            # ==================================================

            build_mode = str(
                lesson.get(
                    "build_mode",
                    "NEW"
                )
            ).strip().upper()
            publication_mode = str(
                lesson.get(
                    "publication_mode",
                    "IMMEDIATE"
                )
            ).strip().upper()
            
            if publication_mode not in (
                    "IMMEDIATE",
                    "GENERATE_ONLY"
            ):
                raise RuntimeError(
                    "Invalid publication_mode for "
                    f"{lesson_package_id}: "
                    f"{publication_mode}"
                )
            update_components = (
                lesson.get(
                    "update_components",
                    []
                )
                or []
            )

            print("=" * 60)
            print("PIPELINE EXECUTION MODE")
            print("Build Mode:", build_mode)
            print(
                "Publication Mode:",
                publication_mode
            )
            print(
                "Update Components:",
                update_components
                or "ALL - NEW BUILD"
            )
            print("=" * 60)

            # ==================================================
            # Prompt Selection
            # ==================================================

            if build_mode == "NEW":

                prompt_sequence = [

                    "LESSON_CONTENT",
                    "DISPLAY_TITLE",
                    "MISSION",

                    "GAMMA_SLIDES",
                    "DID_YOU_KNOW",

                    "QUIZ",
                    "CHECKING_YOUR_THINKING",

                    "ACTIVITIES",
                    "LETS_DO_IT",

                    "RECAP",
                    "WHAT_WE_DISCOVERED",
                    "IMAGE"

                ]

            else:

                prompt_map = {

                    "lesson_content": [
                        "LESSON_CONTENT",
                        "DISPLAY_TITLE",
                        "MISSION"
                    ],

                    "slides": [
                        "GAMMA_SLIDES",
                        "DID_YOU_KNOW"
                    ],

                    "quiz": [
                        "QUIZ",
                        "CHECKING_YOUR_THINKING"
                    ],

                    "activities": [
                        "ACTIVITIES",
                        "LETS_DO_IT"
                    ],

                    "recap": [
                        "RECAP",
                        "WHAT_WE_DISCOVERED"
                    ],

                    "image": [
                        "IMAGE"
                    ]

                }

                prompt_sequence = []

                for component in update_components:

                    prompt_sequence.extend(
                        prompt_map.get(
                            component,
                            []
                        )
                    )

                prompt_sequence = list(
                    dict.fromkeys(
                        prompt_sequence
                    )
                )

            print(
                "Prompt Sequence:",
                prompt_sequence
            )

            prompt_results = []

            self._report_progress(
                progress_url,
                progress_job_id,
                "PROMPTS",
                (
                    "Generating selected lesson prompts..."
                    if build_mode == "UPDATE"
                    else
                    "Generating lesson prompts..."
                ),
                30
            )

            for prompt_type in prompt_sequence:

                print("CALLING PROMPT ENGINE")
                print("PROMPT TYPE:", prompt_type)
                print("LESSON PACKAGE:", lesson_package_id)
                print("=" * 60)

                response = requests.post(
                    PROMPT_ENGINE_URL,
                    json={
                        "workbook_path": str(build["path"]),
                        "lesson_package_id": lesson_package_id,
                        "prompt_type": prompt_type
                    },
                    timeout=PROMPT_TIMEOUT)
                print("PROMPT:", prompt_type)
                print("STATUS :", response.status_code)

                if response.status_code != 200:
                    print(response.text)
                response.raise_for_status()

                result = response.json()

                print("=" * 60)
                print("PROMPT RESULT")
                print(result)
                print("=" * 60)

                prompt_results.append(result)

            print("=" * 60)
            print("PROMPT GENERATION COMPLETE")
            print(
                "Prompts Executed:",
                prompt_sequence
            )
            print("=" * 60)
            engine_build_root = str(build["path"].parent.parent)
            engine_build_name = build["path"].stem

            if build_mode == "NEW":

                engines = [

                    (
                        "Gamma",
                        GAMMA_ENGINE_URL
                    ),

                    (
                        "Quiz",
                        QUIZ_ENGINE_URL
                    ),

                    (
                        "Activities",
                        ACTIVITIES_ENGINE_URL
                    ),

                    (
                        "Recap",
                        RECAP_ENGINE_URL
                    ),

                    (
                       "Image",
                       IMAGE_ENGINE_URL
                    )

                ]

            else:

                engine_map = {

                    "slides": (
                        "Gamma",
                        GAMMA_ENGINE_URL
                    ),

                    "quiz": (
                        "Quiz",
                        QUIZ_ENGINE_URL
                    ),

                    "activities": (
                        "Activities",
                        ACTIVITIES_ENGINE_URL
                    ),

                    "recap": (
                        "Recap",
                        RECAP_ENGINE_URL
                    ),

                    "image": (
                        "Image",
                        IMAGE_ENGINE_URL
                    )

                }

                engines = []

                for component in update_components:

                    engine = engine_map.get(
                        component
                    )

                    if engine:

                        engines.append(
                            engine
                        )

            print(
                "Engines Selected:",
                [
                    name
                    for name, url in engines
                ]
            )
            
            engine_progress = {
                "Gamma": (
                    "SLIDES",
                    "Generating Did You Know slides...",
                    45
                ),
                "Quiz": (
                    "QUIZ",
                    "Generating quiz...",
                    55
                ),
                "Activities": (
                    "ACTIVITIES",
                    "Generating activities...",
                    65
                ),
                "Recap": (
                    "RECAP",
                    "Generating lesson recap...",
                    75
                ),
                "Image": (
                    "IMAGE",
                    "Generating lesson elaboration image...",
                    80
                ),
            }

            for engine_name, engine_url in engines:
                (
                    progress_stage,
                    progress_message,
                    progress_percent
                ) = engine_progress[
                    engine_name
                ]

                self._report_progress(
                    progress_url,
                    progress_job_id,
                    progress_stage,
                    progress_message,
                    progress_percent
                )

                print("=" * 60)
                print(f"CALLING {engine_name.upper()} ENGINE")
                print("=" * 60)
                

                engine_payload = {

                    "build_root":
                        engine_build_root,

                    "build_name":
                        engine_build_name,

                    "lesson_package_id":
                        lesson_package_id

                }

                if engine_name == "Image":

                    engine_payload[
                        "parent_code"
                    ] = lesson.get(
                        "parent_code",
                        ""
                    )

                    engine_payload[
                        "curriculum_code"
                    ] = lesson.get(
                        "curriculum_code",
                        ""
                    )

                    engine_payload[
                        "elaboration"
                    ] = lesson.get(
                        "elaboration",
                        ""
                    )

                    engine_payload[
                        "force_regenerate"
                    ] = (
                        build_mode == "UPDATE"
                    )

                response = requests.post(

                    engine_url,

                    json=engine_payload,

                    timeout=ENGINE_TIMEOUT

                )

                print(engine_name, response.status_code)

                if response.status_code != 200:
                    print(response.text)
                response.raise_for_status()

            # ==================================================
            # Selective UPDATE Moodle Publication
            # ==================================================

            if build_mode == "UPDATE":

                print("=" * 60)
                print("UPDATE GENERATION COMPLETE")
                print(
                    "Components:",
                    update_components
                )
                print("=" * 60)

                # ==============================================
                # Previous Published Moodle Identity
                # ==============================================

                published_build = get_published_build(
                    elaboration_key
                )

                if not published_build:

                    raise RuntimeError(
                        "UPDATE requires a previous "
                        "PUBLISHED registry record."
                    )

                print("=" * 60)
                print("PREVIOUS PUBLISHED BUILD")
                print(
                    "Build:",
                    published_build.get(
                        "build_id"
                    )
                )
                print(
                    "Course:",
                    published_build.get(
                        "moodle_course_id"
                    )
                )
                print(
                    "Recap CMID:",
                    published_build.get(
                        "moodle_recap_cmid"
                    )
                )
                print("=" * 60)

                # ==============================================
                # Inherit Stable Moodle Identity
                # ==============================================

                inherit_moodle_identity(
                    registry_record_id,
                    published_build
                )

                mark_update_ready(
                    registry_record_id,
                    update_components=",".join(
                        update_components
                    )
                )

                print(
                    "REGISTRY RECORD:",
                    registry_record_id,
                    "STATUS: UPDATE_READY"
                )

                # ==============================================
                # Publisher UPDATE
                # ==============================================

                publisher_update_url = (
                    PUBLISHER_ENGINE_URL.rstrip("/")
                    + "/update"
                )

                self._report_progress(
                    progress_url,
                    progress_job_id,
                    "MOODLE_UPDATE",
                    "Updating selected components in Moodle...",
                    85
                )

                print("=" * 60)
                print("CALLING PUBLISHER UPDATE")
                print(
                    "URL:",
                    publisher_update_url
                )
                print("=" * 60)

                response = requests.post(
                    publisher_update_url,
                    json={
                        "build_root":
                            engine_build_root,

                        "build_name":
                            engine_build_name,

                        "lesson_package_id":
                            lesson_package_id,

                        "update_components":
                            update_components,

                         "moodle_identity":
                            {
                                "moodle_course_id":
                                    published_build.get(
                                        "moodle_course_id"
                                    ),

                                "moodle_content_description_cmid":
                                    published_build.get(
                                        "moodle_content_description_cmid"
                                    ),

                                "moodle_lesson_content_cmid":
                                    published_build.get(
                                        "moodle_lesson_content_cmid"
                                    ),

                                "moodle_did_you_know_cmid":
                                    published_build.get(
                                        "moodle_did_you_know_cmid"
                                    ),

                                "moodle_activities_cmid":
                                    published_build.get(
                                        "moodle_activities_cmid"
                                    ),

                                "moodle_recap_cmid":
                                    published_build.get(
                                        "moodle_recap_cmid"
                                    ),
                            }
                    },
                    timeout=ENGINE_TIMEOUT
                )

                print(
                    "Publisher UPDATE:",
                    response.status_code
                )

                if response.status_code != 200:

                    print(response.text)

                response.raise_for_status()

                publisher_result = response.json()

                if (
                    publisher_result.get("status")
                    != "SUCCESS"
                ):

                    raise RuntimeError(
                        "Publisher UPDATE did not "
                        "return SUCCESS for "
                        f"{lesson_package_id}: "
                        f"{publisher_result}"
                    )

                # ==============================================
                # Mark New Build Published
                # ==============================================

                self._report_progress(
                    progress_url,
                    progress_job_id,
                    "REGISTRY",
                    "Recording updated Moodle publication...",
                    95
                )

                mark_published(
                    record_id=
                        registry_record_id,

                    moodle_course_id=
                        published_build.get(
                            "moodle_course_id"
                        ),

                    moodle_section_id=
                        published_build.get(
                            "moodle_section_id"
                        ),

                    moodle_subsection_cmid=
                        published_build.get(
                            "moodle_subsection_cmid"
                        ),

                    moodle_subsection_section_id=
                        published_build.get(
                            "moodle_subsection_section_id"
                        ),

                    moodle_content_description_cmid=
                        published_build.get(
                            "moodle_content_description_cmid"
                        ),

                    moodle_lesson_content_cmid=
                        published_build.get(
                            "moodle_lesson_content_cmid"
                        ),

                    moodle_did_you_know_cmid=
                        published_build.get(
                            "moodle_did_you_know_cmid"
                        ),

                    moodle_quiz_id=
                        published_build.get(
                            "moodle_quiz_id"
                        ),

                    moodle_quiz_cmid=
                        published_build.get(
                            "moodle_quiz_cmid"
                        ),

                    moodle_activities_cmid=
                        published_build.get(
                            "moodle_activities_cmid"
                        ),

                    moodle_recap_cmid=
                        published_build.get(
                            "moodle_recap_cmid"
                        ),

                    update_components=
                        ",".join(
                            update_components
                        )
                )

                print("=" * 60)
                print("UPDATE PUBLISHED SUCCESSFULLY")
                print(
                    "REGISTRY RECORD:",
                    registry_record_id
                )
                print(
                    "Lesson:",
                    lesson_package_id
                )
                print(
                    "Components:",
                    update_components
                )
                print("=" * 60)

                results.append({
                    "lesson_package_id":
                        lesson_package_id,

                    "status":
                        "SUCCESS",

                    "build_mode":
                        "UPDATE",

                    "update_components":
                        update_components,

                    "prompts":
                        prompt_results,

                    "publisher":
                        publisher_result
                })

                continue
            
            # ==================================================
            # Generate Only - Skip Moodle Publication
            # ==================================================

            
            if publication_mode == "GENERATE_ONLY":
                mark_generated(
                    registry_record_id
                )

                print(
                    "REGISTRY RECORD:",
                    registry_record_id,
                    "STATUS: GENERATED"
                )

                self._set_publication_status(
                    build["path"],
                    lesson_package_id,
                    "PENDING",
                    "YES"
                )

                print("=" * 60)
                print("MOODLE PUBLICATION SKIPPED")
                print("Lesson :", lesson_package_id)
                print("Reason : GENERATE_ONLY")
                print("=" * 60)

                publisher_result = {
                    "status": "SKIPPED",
                    "reason": "GENERATE_ONLY"
                }

                results.append({
                    "lesson_package_id":
                        lesson_package_id,

                    "status":
                        "SUCCESS",

                    "prompts":
                        prompt_results,

                    "publisher":
                        publisher_result
                })

                print("=" * 60)
                print("LESSON PACKAGE BUILD COMPLETED")
                print("Build ID :", build["build_id"])
                print("Lesson   :", lesson_package_id)
                print("Publication: PENDING")
                print("=" * 60)

                continue

            # ==================================================
            # Moodle Publication
            # ==================================================

            publication_status = self._get_publication_status(
                build["path"],
                lesson_package_id
            )

            print("=" * 60)
            print("MOODLE PUBLICATION STATUS")
            print("Lesson :", lesson_package_id)
            print("Status :", publication_status or "NOT PUBLISHED")
            print("=" * 60)

            publisher_result = None
            moodle_result = None
            lesson_result_status = "SUCCESS"

            if publication_status == "PUBLISHED":

                print("=" * 60)
                print("PUBLISHER SKIPPED")
                print(
                    lesson_package_id,
                    "is already PUBLISHED."
                )
                print("=" * 60)

                publisher_result = {
                    "status": "SKIPPED",
                    "reason": "ALREADY_PUBLISHED"
                }
                
                                #
                # Workbook reports this lesson as already
                # published. Never mark the new registry
                # record PUBLISHED without Moodle identity.
                #

                previous_published = (
                    get_previous_published_build(
                        elaboration_key,
                        registry_record_id
                    )
                )

                if not previous_published:

                    raise RuntimeError(
                        "Workbook reports PUBLISHED for "
                        f"{lesson_package_id}, but no earlier "
                        "PUBLISHED registry record with Moodle "
                        "identity exists. Refusing to create a "
                        "PUBLISHED registry record without "
                        "Moodle CMIDs."
                    )

                inherit_moodle_identity(
                    registry_record_id,
                    previous_published
                )

                mark_published(
                    record_id=
                        registry_record_id,

                    moodle_course_id=
                        previous_published.get(
                            "moodle_course_id"
                        ),

                    moodle_section_id=
                        previous_published.get(
                            "moodle_section_id"
                        ),

                    moodle_subsection_cmid=
                        previous_published.get(
                            "moodle_subsection_cmid"
                        ),

                    moodle_subsection_section_id=
                        previous_published.get(
                            "moodle_subsection_section_id"
                        ),

                    moodle_content_description_cmid=
                        previous_published.get(
                            "moodle_content_description_cmid"
                        ),

                    moodle_lesson_content_cmid=
                        previous_published.get(
                            "moodle_lesson_content_cmid"
                        ),

                    moodle_did_you_know_cmid=
                        previous_published.get(
                            "moodle_did_you_know_cmid"
                        ),

                    moodle_quiz_id=
                        previous_published.get(
                            "moodle_quiz_id"
                        ),

                    moodle_quiz_cmid=
                        previous_published.get(
                            "moodle_quiz_cmid"
                        ),

                    moodle_activities_cmid=
                        previous_published.get(
                            "moodle_activities_cmid"
                        ),

                    moodle_recap_cmid=
                        previous_published.get(
                            "moodle_recap_cmid"
                        )
                )

                print(
                    "REGISTRY RECORD:",
                    registry_record_id,
                    "STATUS: PUBLISHED"
                )

            else:
                self._report_progress(
                    progress_url,
                    progress_job_id,
                    "MOODLE_PUBLISH",
                    "Publishing lesson to Moodle...",
                    85
                )

                print("=" * 60)
                print("CALLING PUBLISHER ENGINE")
                print("=" * 60)
                self._set_publication_status(
                    build["path"],
                    lesson_package_id,
                    "PUBLISHING",
                    "YES"
                )
                publish_call = (
                    self._call_publisher_safely(
                        engine_build_root,
                        engine_build_name,
                        lesson_package_id
                    )
                )

                publisher_result = (
                    publish_call["result"]
                )

                publication_succeeded = (
                    publish_call["success"]
                )

                if publication_succeeded:
                    moodle_result = publisher_result.get(
                        "publisher"
                    )

                    if (
                        not isinstance(moodle_result, dict)
                        or moodle_result.get("status") != "success"
                    ):
                        publication_succeeded = False
                        publisher_result = {
                            "status": "FAILED",
                            "reason":
                                "INVALID_MOODLE_PUBLISHER_RESULT",
                            "error":
                                "Missing successful Moodle result."
                        }

                if not publication_succeeded:

                    self._set_publication_status(
                        build["path"],
                        lesson_package_id,
                        "FAILED",
                        "YES"
                    )

                    mark_failed(
                        registry_record_id
                    )

                    lesson_result_status = "FAILED"

                    self._report_progress(
                        progress_url,
                        progress_job_id,
                        "MOODLE_PUBLISH_FAILED",
                        "Content generated, but Moodle publication failed.",
                        95
                    )

                else:

                    self._report_progress(
                        progress_url,
                        progress_job_id,
                        "REGISTRY",
                        "Recording Moodle publication...",
                        95
                    )

                    mark_published(
                        record_id=
                            registry_record_id,

                        moodle_course_id=
                            moodle_result.get(
                                "courseid"
                            ),

                        moodle_section_id=
                            moodle_result.get(
                                "strandsectionid"
                            ),

                        moodle_subsection_cmid=
                            moodle_result.get(
                                "subsectioncmid"
                            ),

                        moodle_subsection_section_id=
                            moodle_result.get(
                                "subsectionsectionid"
                            ),

                        moodle_content_description_cmid=
                            moodle_result.get(
                                "contentdescriptioncmid"
                            ),

                        moodle_lesson_content_cmid=
                            moodle_result.get(
                                "lessoncontentcmid"
                            ),

                        moodle_did_you_know_cmid=
                            moodle_result.get(
                                "didyouknowcmid"
                            ),

                        moodle_quiz_id=
                            moodle_result.get(
                                "quizid"
                            ),

                        moodle_quiz_cmid=
                            moodle_result.get(
                                "quizcmid"
                            ),

                        moodle_activities_cmid=
                            moodle_result.get(
                                "activitiescmid"
                            ),

                        moodle_recap_cmid=
                            moodle_result.get(
                                "recapcmid"
                            )
                    )

                    published_questions = (
                        moodle_result.get(
                            "questions"
                        )
                    )

                    if not published_questions:
                        raise RuntimeError(
                            "Successful NEW Moodle publication "
                            "did not return question mappings."
                        )

                    register_quiz_questions(
                        build_id=
                            build["build_id"],

                        lesson_package_id=
                            lesson_package_id,

                        curriculum_code=
                            lesson.get(
                                "curriculum_code"
                            ),

                        moodle_course_id=
                            moodle_result.get(
                                "courseid"
                            ),

                        moodle_quiz_id=
                            moodle_result.get(
                                "quizid"
                            ),

                        moodle_quiz_cmid=
                            moodle_result.get(
                                "quizcmid"
                            ),

                        questions=
                            published_questions,

                        source=
                            "PUBLISH"
                    )

                    print(
                        "QUIZ QUESTIONS REGISTERED:",
                        len(
                            published_questions
                        )
                    )
                    print(
                        "REGISTRY RECORD:",
                        registry_record_id,
                        "STATUS: PUBLISHED"
                    )
            #
            # Finished
            #

            print("=" * 60)
            print("LESSON PACKAGE BUILD COMPLETED")
            print("Build ID :", build["build_id"])
            print("Lesson   :", lesson_package_id)
            print("=" * 60)

            results.append({
                "lesson_package_id":
                    lesson_package_id,

                "status":
                    lesson_result_status,

                "prompts":
                    prompt_results,

                "publisher":
                    publisher_result
            })

        failed_lessons = [
            item
            for item in results
            if item.get("status") == "FAILED"
        ]

        return {

            "status": (
                "PARTIAL_FAILURE"
                if failed_lessons
                else "SUCCESS"
            ),

            "build_id": build["build_id"],

            "filename": build["filename"],

            "workbook_path": str(build["path"]),

            "lessons": results

        }
