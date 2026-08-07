from openpyxl import load_workbook

from config import (
    SHEET_BUILD_METADATA,
    SHEET_LESSON_DB,
    SHEET_PROMPT_QUEUE
)


# ==========================================================
# Workbook Reader
# ==========================================================

class WorkbookReader:

    def __init__(

            self,

            workbook_path

    ):

        self.workbook_path = workbook_path

        self.workbook = load_workbook(

            workbook_path,

            data_only=True

        )

    # ======================================================
    # Header Map
    # ======================================================

    @staticmethod
    def header_map(sheet):

        headers = {}

        for cell in sheet[1]:

            if cell.value:
                headers[

                    str(cell.value).strip()

                ] = cell.column

        return headers

    # ======================================================
    # Build Metadata
    # ======================================================

    def build_metadata(self):

        sheet = self.workbook[

            SHEET_BUILD_METADATA

        ]

        headers = self.header_map(sheet)

        metadata = {}

        for field, column in headers.items():
            metadata[field] = sheet.cell(

                row=2,

                column=column

            ).value

        return metadata

    # ======================================================
    # Lesson DB
    # ======================================================

    def lessons(self):

        sheet = self.workbook[

            SHEET_LESSON_DB

        ]

        headers = self.header_map(sheet)

        lessons = []

        row = 2

        while True:

            lesson_package_id = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not lesson_package_id:
                break

            lesson = {}

            for field, column in headers.items():
                lesson[field] = sheet.cell(

                    row=row,

                    column=column

                ).value

            lessons.append(

                lesson

            )

            row += 1

        return lessons

    # ======================================================
    # Prompt Queue
    # ======================================================

    def prompt_queue(self):

        sheet = self.workbook[

            SHEET_PROMPT_QUEUE

        ]

        headers = self.header_map(sheet)

        queue = []

        row = 2

        while True:

            prompt_id = sheet.cell(

                row=row,

                column=headers["prompt_id"]

            ).value

            if not prompt_id:
                break

            record = {}

            for field, column in headers.items():
                record[field] = sheet.cell(

                    row=row,

                    column=column

                ).value

            queue.append(

                record

            )

            row += 1

        return queue

    # ======================================================
    # One Lesson
    # ======================================================

    def lesson(

            self,

            lesson_package_id

    ):

        for lesson in self.lessons():

            if (

                    lesson["lesson_package_id"]

                    ==

                    lesson_package_id

            ):
                return lesson

        return None

    # ======================================================
    # One Prompt
    # ======================================================

    def prompt(

            self,

            prompt_id

    ):

        for record in self.prompt_queue():

            if (

                    record["prompt_id"]

                    ==

                    prompt_id

            ):
                return record

        return None

    # ======================================================
    # Close
    # ======================================================

    def close(self):

        self.workbook.close()
