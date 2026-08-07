from pathlib import Path

from openpyxl import load_workbook


# ==========================================================
# Workbook Updater
# ==========================================================

class WorkbookUpdater:

    def __init__(

            self,

            workbook_path

    ):

        self.workbook_path = Path(

            workbook_path

        )

        self.workbook = load_workbook(

            self.workbook_path

        )

    # ======================================================
    # Header Map
    # ======================================================

    @staticmethod
    def header_map(sheet):

        headers = {}

        for cell in sheet[1]:

            if cell.value:
                headers[
                    str(cell.value).strip()
                ] = cell.column

        return headers

    # ======================================================
    # Find Lesson
    # ======================================================

    def find_row(

            self,

            sheet_name,

            lesson_package_id

    ):

        sheet = self.workbook[

            sheet_name

        ]

        headers = self.header_map(sheet)

        row = 2

        while True:

            value = sheet.cell(

                row=row,

                column=headers["lesson_package_id"]

            ).value

            if not value:
                return None

            if value == lesson_package_id:
                return row

            row += 1

    # ======================================================
    # Update Fields
    # ======================================================

    def update(

            self,

            sheet_name,

            lesson_package_id,

            values

    ):

        sheet = self.workbook[

            sheet_name

        ]

        headers = self.header_map(sheet)

        row = self.find_row(

            sheet_name,

            lesson_package_id

        )

        if row is None:
            raise ValueError(

                f"{lesson_package_id} not found"

            )

        for field, value in values.items():

            if field not in headers:
                continue

            sheet.cell(

                row=row,

                column=headers[field]

            ).value = value

    # ======================================================
    # Append Row
    # ======================================================

    def append(

            self,

            sheet_name,

            values

    ):

        sheet = self.workbook[

            sheet_name

        ]

        sheet.append(values)

    # ======================================================
    # Save
    # ======================================================

    def save(self):

        self.workbook.save(

            self.workbook_path

        )

        self.workbook.close()
