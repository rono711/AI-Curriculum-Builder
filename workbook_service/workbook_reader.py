from pathlib import Path

from openpyxl import load_workbook
# from shared.workbook.headers import normalize_header

# ==========================================================
# Workbook Reader
# ==========================================================

class WorkbookReader:

    def __init__(

            self,

            workbook_path

    ):

        self.workbook_path = Path(

            workbook_path

        )

        self.workbook = load_workbook(

            self.workbook_path,

            data_only=True

        )

    # ======================================================
    # Worksheet
    # ======================================================

    def worksheet(

            self,

            sheet_name

    ):

        return self.workbook[sheet_name]

    # ======================================================
    # Header Map
    # ======================================================

    def header_map(self,sheet):

        headers = {}

        for cell in sheet[1]:

            if cell.value:
                headers[
                    str(cell.value).strip()
                ] = cell.column

        return headers

    # ======================================================
    # Read Worksheet
    # ======================================================

    def rows(

            self,

            sheet_name

    ):

        sheet = self.worksheet(

            sheet_name

        )

        headers = self.header_map(sheet)

        rows = []

        row_number = 2

        while True:

            first = sheet.cell(

                row=row_number,

                column=1

            ).value

            if first in (

                    None,

                    ""

            ):
                break

            record = {}

            for field, column in headers.items():
                record[field] = sheet.cell(

                    row=row_number,

                    column=column

                ).value

            rows.append(

                record

            )

            row_number += 1

        return rows

    # ======================================================
    # Find One Record
    # ======================================================

    def record(

            self,

            sheet_name,

            key_field,

            key_value

    ):

        for row in self.rows(

                sheet_name

        ):

            if row.get(

                    key_field

            ) == key_value:
                return row

        return None

    # ======================================================
    # Build Metadata
    # ======================================================

    def metadata(self):

        sheet = self.worksheet(

            "Build_Metadata"

        )

        headers = self.header_map(sheet)

        metadata = {}

        for field, column in headers.items():
            metadata[field] = sheet.cell(

                row=2,

                column=column

            ).value

        return metadata

    # ======================================================
    # Lesson
    # ======================================================
    print("=" * 60)
    print("METADATA")
    print(metadata)
    print("=" * 60)
    def lesson(

            self,

            lesson_package_id

    ):

        return self.record(

            "Lesson_DB",

            "lesson_package_id",

            lesson_package_id

        )

    # ======================================================
    # Prompt
    # ======================================================

    def prompt(

            self,

            prompt_id

    ):

        return self.record(

            "Prompt_Queue",

            "prompt_id",

            prompt_id

        )

    # ======================================================
    # AI Job
    # ======================================================

    def ai_job(

            self,

            job_id

    ):

        return self.record(

            "AI_Generation",

            "job_id",

            job_id

        )

    # ======================================================
    # Close
    # ======================================================

    def close(self):

        self.workbook.close()
